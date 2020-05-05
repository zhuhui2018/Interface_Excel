#encoding:utf-8

from openpyxl import load_workbook
from openpyxl.styles import Border,Side,Font
import time

class ParseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None) #设置字体的颜色
        #颜色对应的RGB值
        self.RGBDict = {'red': 'FFFF3030', 'green': 'FF008B00'}

    def loadWorkBook(self,excelPathAndName):
        #将文件加载到内存，并获取workbook对象
        try:
            self.workbook = load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self,sheetName):
        #根据sheet名字获取该sheet对象
        try:
            sheet = self.workbook[sheetName]
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self,sheetIndex):
        #根据sheet的索引号获取sheet对象
        try:
            sheetname = self.workbook.sheetnames[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook[sheetname]
        return sheet

    def getRowsNumber(self,sheet):
        #获取sheet中有数据区域的结束行号
        return sheet.max_row

    def getColsNumber(self,sheet):
        #获取sheet中有数据区域的结束列号
        return sheet.max_column

    def getStartRowNumber(self,sheet):
        #获取sheet中有数据区域的开始的行号
        return sheet.min_row

    def getStartColNumber(self,sheet):
        #获取sheet中有数据区域的开始的列号
        return sheet.min_column

    def getRow(self,sheet,rowNo):
        #获取sheet中某一行，返回的是这一行所有的数据内容组成的tuple
        try:
            rows=[]
            for row in sheet.iter_rows():
                rows.append(row)
            return rows[rowNo-1]
        except Exception as e:
            raise e

    def getColumn(self,sheet,colNo):
        #获取sheet中某一列，返回的是这一列所有数据内容组成的tuple
        try:
            cols = []
            for col in sheet.iter_cols():
                cols.append(col)
            return cols[colNo-1]
        except Exception as e:
            raise e

    def getCellOfValue(self,sheet,coordinate=None,
                       rowNo=None,colNo=None):
        #获取单元格所在的位置索引获取单元格中的值
        if coordinate !=None:
            try:
                return sheet[coordinate]
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None \
            and colNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")

    def getCellOfObject(self, sheet, coordinate = None,
                        rowNo = None, colsNo = None):
        # 获取某个单元格的对象，可以根据单元格所在位置的数字索引，
        # 也可以直接根据excel中单元格的编码及坐标
        # 如getCellObject(sheet, coordinate = 'A1') or
        # getCellObject(sheet, rowNo = 1, colsNo = 2)
        if coordinate != None:
            try:
                # return sheet.cell(coordinate = coordinate)
                return sheet[coordinate]
            except Exception as err:
                raise err
        elif coordinate == None and rowNo is not None and \
                        colsNo is not None:
            try:
                return sheet.cell(row = rowNo,column = colsNo)
            except Exception as err:
                raise err
        else:
            raise Exception("Insufficient Coordinates of cell !")

    def writeCell(self, sheet, content, coordinate = None,
        rowNo = None, colsNo = None, style = None):
        #根据单元格在excel中的编码坐标或者数字索引坐标向单元格中写入数据，
        # 下标从1开始，参style表示字体的颜色的名字,比如red，green
        if coordinate is not None:
            try:
                # sheet.cell(coordinate = coordinate).value = content
                sheet[coordinate] = content
                if style is not None:
                    sheet[coordinate].\
                        font = Font(color = self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and \
                        colsNo is not None:
            try:
                sheet.cell(row = rowNo,column = colsNo).value = content
                if style:
                    sheet.cell(row = rowNo,column = colsNo).\
                        font = Font(color = self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")

    def writeCellCurrentTime(self, sheet, coordinate = None,
                rowNo = None, colsNo = None):
        # 写入当前的时间，下标从1开始
        now = int(time.time())  #显示为时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate = coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None \
                and colsNo is not None:
            try:
                sheet.cell(row = rowNo, column = colsNo
                        ).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell !")

if __name__ == "__main__":
    p=ParseExcel()
    p.loadWorkBook(r"d:\126邮箱联系人.xlsx")
    sheetobj = p.getSheetByName("联系人")
    print(p.getStartColNumber(sheetobj))
    p.writeCell(sheetobj,"hello",rowNo=3,colsNo=4)
    p.writeCellCurrentTime(sheetobj,rowNo=6,colsNo=8)